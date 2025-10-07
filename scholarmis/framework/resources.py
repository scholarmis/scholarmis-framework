import os
import uuid
import decimal
import json
import tablib
import io
import traceback
from celery import Task
from io import BytesIO
from django.forms import ValidationError
from django.core.files.storage import default_storage
from slugify import slugify 
from tablib import Dataset
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple
from django.apps import apps
from django.db.models import Field, Model, QuerySet
from django.http import HttpResponse, StreamingHttpResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from import_export.resources import ModelResource
from import_export.widgets import  DateWidget
from import_export.widgets import  ForeignKeyWidget as BaseForeignKeyWidget



class WorksheetHelper:
    """
    Helper class for manipulating worksheets, including styling, data extraction, 
    and protection functionality related to a Django model's fields.
    """

    @staticmethod
    def clean(value):
        """
        Cleans the given value for safe inclusion in an Excel worksheet.

        Args:
            value: The value to be cleaned.

        Returns:
            A cleaned value that can be safely written to an Excel worksheet.
        """
        if isinstance(value, (datetime, date)):
            # Format datetime and date objects
            return value.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(value, uuid.UUID):
            # Convert UUIDs to strings
            return str(value)
        elif isinstance(value, (list, dict, set, tuple)):
            # Convert lists, dictionaries, sets, or tuples to JSON strings
            try:
                return json.dumps(value, default=str)
            except (TypeError, ValueError):
                return str(value)
        elif isinstance(value, decimal.Decimal):
            # Convert Decimal to float for Excel compatibility
            return float(value)
        elif isinstance(value, bytes):
            # Convert bytes to a human-readable string
            try:
                return value.decode("utf-8")
            except UnicodeDecodeError:
                return str(value)
        elif isinstance(value, bool):
            # Represent booleans as "TRUE" or "FALSE"
            return "TRUE" if value else "FALSE"
        elif value is None:
            # Replace None with an empty string
            return ""
        elif isinstance(value, (int, float, str)):
            # Directly return int, float, and string types
            return value
        else:
            # Fallback for unsupported types
            return str(value)

    @staticmethod
    def style_header_row(sheet):
        """
        Apply styles to the header row (model field names).
        
        Args:
            sheet: The worksheet object where the header row styles will be applied.
        """
        sheet.freeze_panes = "A2"  # Freeze the header row to keep it visible while scrolling
        header = sheet[1]  # Get the first row (header)
        for cell in header:
            # Apply font, background color, and alignment to header cells
            cell.font = Font(size=12, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def get_for_model(model: Model):
        """
        Generate a valid sheet name, ensuring it doesn't exceed 31 characters.
        
        Args:
            model: The Django model class for which the sheet name is being generated.

        Returns:
            A string representing a valid sheet name for the model.
        """
        return model._meta.verbose_name_plural.replace(" ", "_")[:31]

    @staticmethod
    def get_foreignkey_dataset(field: Field):
        """
        Generate a dataset for a ForeignKey reference field, including the ID and value.
        
        Args:
            field: The ForeignKey field from the model.

        Returns:
            A Dataset containing IDs and string representations of the related objects.
        """
        if hasattr(field, "_custom_queryset"):
            queryset = getattr(field, '_custom_queryset', None)
        else:
            queryset = field.related_model.objects.all()  # Get all related model instances

        # Create dataset with headers "ID" and "Value"
        dataset = Dataset(headers=["ID", "Value"])
        
        index = 0
        for obj in queryset:
            if isinstance(obj, tuple) and len(obj) >= 2:
                # Assuming obj is a tuple (id, code), where obj[0] is the ID and obj[1] is the code
                dataset.append([obj[0], str(obj[1])])  # Append the ID and code from the tuple
            elif isinstance(obj, list) and len(obj) >= 2:
                # Assuming obj is a list with the first element as the ID and second element as the code
                dataset.append([obj[0], str(obj[1])])  # Append the ID and code from the list
            elif hasattr(obj, "pk"):
                # Handle model instances with a pk attribute
                dataset.append([obj.pk, str(obj)])  # Append the object's PK and string representation
            elif hasattr(obj, "id"):
                # Handle model instances with an id attribute
                dataset.append([obj.id, str(obj)])  # Append the object's ID and string representation
            else:
                # If the object doesn't have a pk or id, use the index
                dataset.append([index, str(obj)])
                index += 1  # Increment the index
        return dataset

    @staticmethod
    def get_choice_dataset(field: Field):
        """
        Generate a dataset for a ChoiceField reference field, including ID and value.

        Args:
            field: The ChoiceField from the model.

        Returns:
            A Dataset containing IDs and values for each choice.
        """
        choices = getattr(field, "custom_choices", field.choices)  # Use custom choices if defined
        dataset = Dataset(headers=["ID", "Value"])
        for choice in choices:
            dataset.append(choice)  # Append each choice as a tuple
        return dataset

    @staticmethod
    def add_headers(worksheet: Worksheet, dataset: Dataset, start=1):
        """
        Add headers to the worksheet starting at the given column index.
        
        Args:
            worksheet: The worksheet to which headers will be added.
            dataset: The dataset containing headers.
            start: The starting column index for adding headers (default is 1).
        """
        for col_index, header in enumerate(dataset.headers, start):
            worksheet.cell(row=1, column=col_index, value=header)  # Add headers to the first row

    @staticmethod
    def add_rows(worksheet: Worksheet, dataset: Dataset, start=2):
        """
        Add rows of data to the worksheet starting at the given row index.
        
        Args:
            worksheet: The worksheet to which rows will be added.
            dataset: The dataset containing rows of data.
            start: The starting row index for adding data (default is 2).
        """
        for row_index, row in enumerate(dataset.dict, start):
            for col_index, (key, value) in enumerate(row.items(), start=1):
                value = WorksheetHelper.clean(value)
                worksheet.cell(row=row_index, column=col_index, value=value)  # Add data to worksheet cells

    @staticmethod
    def convert_uuid_to_string(value):
        """Convert UUID fields in the DataFrame to strings."""
        if isinstance(value, uuid.UUID):
            return str(value)
        return value

   
    @staticmethod
    def protect(worksheet: Worksheet, password):
        """
        Protect the worksheet by hiding it and setting a password.

        Args:
            worksheet: The worksheet to be protected.
            password: The password to lock the worksheet with.
        """
        worksheet.sheet_state = 'hidden'  # Hide the worksheet
        worksheet.protection.sheet = True  # Lock the worksheet
        worksheet.protection.set_password(password)  # Set the protection password


class ReferenceField:
    """
    A utility class to handle Reference fields in a Django model.
    This class is used to manage ForeignKey and ChoiceField relationships.
    """

    @staticmethod
    def get_foreign_fields(model: Model, fields: List[str], foreign_fields: Optional[Dict[str, QuerySet]] = None) -> List[Field]:
        """
        Get ForeignKey fields for a model based on the provided field names. Custom 
        querysets can be passed for the fields.
        
        Args:
            model: The Django model class.
            fields: A list of field names to look for in the model.
            foreign_fields: A dictionary of field names with custom querysets (optional).

        Returns:
            A list of ForeignKey field objects from the model.
        """
        # Get the ForeignKey fields from the model
        model_fk_fields = {
            field.name: field for field in model._meta.get_fields()
            if field.is_relation and field.many_to_one and field.name in fields
        }

        result_fields = []
        foreign_fields = foreign_fields or {}

        for field_name in fields:
            if field_name in foreign_fields:
                # If the field has custom data, attach the custom queryset
                if field_name in model_fk_fields:
                    field = model_fk_fields[field_name]
                    field._custom_queryset = foreign_fields[field_name]  # Set custom queryset
                    result_fields.append(field)
            elif field_name in model_fk_fields:
                # Use the default field if no custom queryset is defined
                result_fields.append(model_fk_fields[field_name])

        return result_fields

    @staticmethod
    def get_choice_fields(model: Model, fields: List[str], choice_fields: Optional[Dict[str, List[tuple]]] = None) -> List[Field]:
        """
        Get ChoiceField fields for a model based on the provided field names. Custom 
        choices can be passed for the fields.
        
        Args:
            model: The Django model class.
            fields: A list of field names to look for in the model.
            choice_fields: A dictionary of field names with custom choices (optional).

        Returns:
            A list of ChoiceField objects from the model.
        """
        # Get the ChoiceField fields from the model
        model_choice_fields = {
            field.name: field for field in model._meta.get_fields()
            if not field.is_relation and field.choices and field.name in fields
        }

        # Apply custom choices if provided
        if choice_fields:
            for field_name, custom_choices in choice_fields.items():
                if field_name in model_choice_fields:
                    # Replace default choices with custom choices
                    field = model_choice_fields[field_name]
                    field.choices = custom_choices
                    model_choice_fields.update({field_name: field})

        return list(model_choice_fields.values())


class ExcelExporter:
    PASSWORD = "password"  # Default password for worksheet protection
    MAX_ROW = 1048576  # Maximum number of rows in an Excel sheet

    def __init__(self, model: Model, dataset: Dataset, reference_sheets: Dict[str, Tuple[str, Dataset]], hidden_fields=[], with_data: bool = False, protect: bool=True, export_name=None):
        """
        Initialize the ExcelExporter instance.
        
        Args:
            model (Model): The Django model being exported.
            dataset (Dataset): The main dataset that will populate the main sheet.
            reference_sheets (Dict[str, Tuple[str, Dataset]]): Reference data for foreign keys or choices, including dataset.
            with_data (bool): Flag to indicate whether to include data in the export (defaults to False).
        """
        self.model = model
        self.dataset = dataset
        self.reference_sheets = reference_sheets
        self.hidden_fields = hidden_fields
        self.with_data = with_data
        self.protect = protect
        self.export_name = export_name

    def build_workbook(self):
        workbook = Workbook()

        # Main sheet
        sheet_name = WorksheetHelper.get_for_model(self.model)
        main_sheet = workbook.active
        main_sheet.title = sheet_name

        # Write headers to the main sheet
        WorksheetHelper.add_headers(main_sheet, self.dataset)

        # Apply styles to the header row
        WorksheetHelper.style_header_row(main_sheet)

        if self.with_data:
            WorksheetHelper.add_rows(main_sheet, self.dataset, 2)

        if len(self.hidden_fields) > 0:
            self._hide_columns(main_sheet, self.hidden_fields)

        # Add additional sheets for foreign keys and choices
        self._add_sheets(workbook, self.reference_sheets)

        # Add data validation for reference fields (dropdowns)
        self._add_validations(main_sheet, self.reference_sheets)

        return workbook
       
    def export(self):
        workbook = self.build_workbook()
        return self._export(workbook)
        
    def stream(self):
        workbook = self.build_workbook()
        return self._stream(workbook)
        
    def save(self, path):
        workbook = self.build_workbook()

        file_name = self._get_file_name(self.model)

        save_path = default_storage.path(path)

        file_path = os.path.join(save_path, file_name)

        workbook.save(file_path)
        return file_path

    def _hide_columns(self, work_sheet: Worksheet, columns):
        """
        Hide columns in an Excel sheet based on the provided headers.

        Args:
            work_sheet (Worksheet): The openpyxl Worksheet object.
            columns (list): A list of headers (column names) to hide.
        """
        # Map headers to their corresponding column indices
        header_row = work_sheet[1]  # Assuming headers are in the first row
        header_map = {cell.value: cell.column for cell in header_row if cell.value}

        # Process each column to hide and lock
        for header in columns:
            if header in header_map:
                col_index = header_map[header]
                col_letter = work_sheet.cell(row=1, column=col_index).column_letter

                # Hide the column
                work_sheet.column_dimensions[col_letter].hidden = True

    def _add_sheets(self, workbook: Workbook, reference_sheets: Dict[str, Tuple[str, Dataset]]):
        """
        Add additional sheets to the workbook for foreign key and choice reference data.
        
        Args:
            workbook (Workbook): The openpyxl workbook to add the sheets to.
            reference_sheets (Dict[str, Tuple[str, Dataset]]): Reference data for foreign keys or choices, including dataset.
        """
        for _, (sheet_name, dataset) in reference_sheets.items():
            reference_sheet = workbook.create_sheet(sheet_name)
            # Add headers and rows to the reference sheet
            WorksheetHelper.add_headers(reference_sheet, dataset)
            WorksheetHelper.add_rows(reference_sheet, dataset, 2)
            # Protect the sheet with a password
            if self.protect:
                WorksheetHelper.protect(reference_sheet, self.PASSWORD)

    def _add_validations(self, main_sheet: Worksheet, reference_sheets: Dict[str, Tuple[str, Dataset]]):
        """
        Add data validation (dropdowns) to the main sheet based on reference data.
        
        Args:
            main_sheet (Worksheet): The main worksheet where validation will be added.
            reference_sheets (Dict[str, Tuple[str, Dataset]]): Reference data for foreign keys or choices.
        """
        for col_index, header in enumerate(self.dataset.headers, start=1):
            column_letter = get_column_letter(col_index)
            field_name = str(header).lower()

            # Check if the column header corresponds to a reference field
            if field_name in reference_sheets:
                sheet_name, reference_sheet = reference_sheets[field_name]
                # Formula for referencing the range in the reference sheet
                formula = f"='{sheet_name}'!$B$2:$B${len(reference_sheet) + 1}"
                # Define data validation for the column
                data_validation = DataValidation(
                    type="list",
                    formula1=formula,
                    allow_blank=True,
                    showDropDown=False
                )
                # Apply the data validation to the whole column
                data_validation.sqref = f"{column_letter}2:{column_letter}{self.MAX_ROW}"
                main_sheet.add_data_validation(data_validation)

    def _export(self, workbook: Workbook):
        """Return the workbook as an HTTP response for download."""
        buffer = BytesIO()
        # Save the workbook to an in-memory buffer
        workbook.save(buffer)
        buffer.seek(0)

        # Generate the file name based on the model's verbose name
        if self.export_name:
            file_name = self._get_export_name()
        else:
            file_name = self._get_file_name(self.model)

        # Create and return the HTTP response with the file attachment
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename={file_name}"
        return response
    
    def _stream(self, workbook: Workbook):
        """Return the workbook as an HTTP response for download."""
        buffer = BytesIO()
        # Save the workbook to an in-memory buffer
        workbook.save(buffer)
        buffer.seek(0)

        # Generate the file name based on the model's verbose name
        if self.export_name:
            file_name = self._get_export_name()
        else:
            file_name = self._get_file_name(self.model)

        # Create and return the HTTP response with the file attachment
        response = StreamingHttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename={file_name}"
        return response
    
    def _get_file_name(self, model: Model):
        """
        Generate a slugified file name based on the model's verbose name.
        
        Args:
            model (Model): The Django model to generate the file name from.
        
        Returns:
            str: The slugified file name for the export file.
        """
        name = str(model._meta.verbose_name_plural)
        return f"{slugify(name, separator='-').upper()}.xlsx"
    
    def _get_export_name(self):
        """
        Generate a slugified file name based on the model's verbose name.
        
        Args:
            model (Model): The Django model to generate the file name from.
        
        Returns:
            str: The slugified file name for the export file.
        """
        name = str(self.export_name)
        return f"{slugify(name, separator='-').upper()}.xlsx"


class ExcelDateWidget(DateWidget):
    def clean(self, value, row=None, *args, **kwargs):
        # Check if the value is a number (Excel serial date format)
        if isinstance(value, (int, float)):
            try:
                # Excel's date system starts from 1900-01-01, and it incorrectly considers 1900 as a leap year
                # Subtract 2 to account for this bug
                excel_base_date = datetime(1900, 1, 1)
                return excel_base_date + timedelta(days=int(value) - 2)
            except Exception as e:
                raise ValidationError(f"Error converting date: {e}")
        
        # If it's already a string, use the default DateWidget behavior
        try:
            return super().clean(value, row, *args, **kwargs)
        except ValueError:
            raise ValidationError(f"Invalid date format: {value}. Expected format: YYYY-MM-DD.")


class ForeignKeyWidget(BaseForeignKeyWidget):

    def clean(self, value, row=None, **kwargs):
        """
        Returns None instead of raising DoesNotExist if value not found.
        """
        if not value:
            return None
        queryset = self.get_queryset(value, row)  # Correct usage
        try:
            return queryset.get(**{self.field: value})
        except self.model.DoesNotExist:
            return None 


class BaseResource(ModelResource):
    """
    A class for exporting data from a Django model into a structured format such as XLSX.
    """

    class Meta:
        model = None  # Must be overridden by subclasses to specify the model to export
        fields = []  # Fields to export
        exclude = []  # Fields to exclude from export
        foreign_fields = {}  # List of ForeignKey fields to include in the export
        choice_fields = {}  # List of fields or dict with custom choices for the export
        hidden_fields = [] # List of fields to include but hide in the export
        with_data = False  # Flag to include data in the export (default False)
        required = True  # Flag to indicate if the resource is required
        skip_unchanged = True
        report_skipped = True
        protect = True
        export_name = None

    def __init__(self, **kwargs):
        """
        Initialize the BaseResource instance with model and other metadata.

        Args:
            **kwargs: Arbitrary keyword arguments for initializing the resource.
        """
        super().__init__(**kwargs)
        # Set model-related metadata
        self.model = self._meta.model
        self.app_label = self.model._meta.app_label
        self.model_name = self.model._meta.model_name
        self.verbose_name = self.model._meta.verbose_name
        self.verbose_name_plural = self.model._meta.verbose_name_plural

        # Set flags for including data and whether the resource is required
        self.with_data = self._meta.with_data
        self.required = self._meta.required
        self.protect = self._meta.protect
        self.hidden_fields = self._meta.hidden_fields
        self.headers = self._meta.fields
        self.export_name = self._meta.export_name

        self.foreignkey_fields = ReferenceField.get_foreign_fields(self.model, self.fields, self._meta.foreign_fields)
        self.choice_fields = ReferenceField.get_choice_fields(self.model, self.fields, self._meta.choice_fields)


    def __str__(self):
        return self.model_name
    
    @property
    def name(self):
        return self.model_name
    
    @property
    def display_name(self):
        return str(self.verbose_name_plural).capitalize()
    
    def set_export_name(self, name):
        self.export_name = name

    
    def get_exporter(self, queryset=None, *args, **kwargs):
        # Prepare the main dataset
        if self.with_data:
            dataset = super().export(queryset, *args, **kwargs)
        else:
            dataset = Dataset()
            dataset.headers = self.headers

        # Create reference sheets (additional sheets for ForeignKey and Choice data)
        reference_sheets = self._create_reference_sheets()
        
        # Export the data to an XLSX file
        return ExcelExporter(
            model=self.model, 
            dataset=dataset, 
            reference_sheets=reference_sheets, 
            hidden_fields=self.hidden_fields,
            with_data=self.with_data, 
            protect=self.protect,
            export_name=self.export_name
        )

    def export(self, queryset=None, *args, **kwargs):
        exporter = self.get_exporter(queryset)
        return exporter.export()
    
    def save_file(self, save_path="exports", queryset=None, *args, **kwargs):
        exporter = self.get_exporter(queryset)
        return exporter.save(save_path)
    
    def _create_reference_sheets(self):
        """
        Create additional sheets by combining ForeignKey reference fields and Choice fields.
        """
        foreignkey_sheets = self._create_foreignkey_sheets()
        choice_sheets = self._create_choice_sheets()
        reference_sheets = {**foreignkey_sheets, **choice_sheets}
        return reference_sheets

    def _create_foreignkey_sheets(self):
        """
        Create additional sheets for ForeignKey reference fields.

        Args:
            foreignkey_fields (List[Field]): A list of ForeignKey fields to process.

        Returns:
            Dict[str, Tuple[str, Dataset]]: A dictionary where the key is the field name and the value is a tuple
                                           containing the sheet name and dataset for the foreign key.
        """
        reference_sheets = {}
        for field in self.foreignkey_fields:
            # Get the dataset for each ForeignKey field
            dataset = WorksheetHelper.get_foreignkey_dataset(field)
            if dataset:
                # Generate the sheet name from the related model's name
                sheet_name = WorksheetHelper.get_for_model(field.remote_field.model)
                reference_sheets[field.name.lower()] = (sheet_name, dataset)
        return reference_sheets

    def _create_choice_sheets(self):
        """
        Create additional sheets for choice fields.

        Args:
            choice_fields (List[Field]): A list of choice fields to process.

        Returns:
            Dict[str, Tuple[str, Dataset]]: A dictionary where the key is the field name and the value is a tuple
                                           containing the sheet name and dataset for the choice field.
        """
        reference_sheets = {}
        for field in self.choice_fields:
            # Get the dataset for each choice field
            dataset = WorksheetHelper.get_choice_dataset(field)
            if dataset:
                # Generate the sheet name from the field name (limited to 31 characters)
                sheet_name = field.name[:31]
                reference_sheets[field.name.lower()] = (sheet_name, dataset)
        return reference_sheets

    def is_blank(self):
        """
        Check if the model has no records.

        Returns:
            bool: True if the model has no records, False otherwise.
        """
        model = apps.get_model(self.app_label, self.model_name)
        return model.objects.count() == 0

    def is_filled(self):
        """
        Check if the model has records.

        Returns:
            bool: True if the model has records, False if it is empty.
        """
        return not self.is_blank()

    def import_file(self,file, format="xlsx", raise_errors=False):
        dataset = tablib.Dataset().load(file, format)
        result = self.import_data(dataset=dataset, raise_errors=raise_errors)
        return result


class ResourceImport:

    def __init__(self, task: Task, resource: BaseResource, file_path: str, user, raise_errors=True):
        """
        Initialize ResourceImport with resource instance, file path, and user.

        :param resource: The resource instance to handle data import.
        :param file_path: Path to the file to be processed.
        :param user: The user who initiated the task.
        """
        self.task = task
        self.resource = resource
        self.file_path = file_path
        self.user = user
        self.raise_errors = raise_errors
        self.file_content, self.file_format = self.read_excel_file(file_path)  # Read file content
        self.dataset = tablib.Dataset().load(self.file_content, self.file_format)  # Load the dataset

        self.dataset = self.filter(self.dataset)
        self.headers = self.dataset.headers
        self.notifier: callable = None
    
    def set_notifier(self, notifier: callable):
        self.notifier = notifier

    def read_excel_file(self, file_path, size=None):
        # Get the file extension
        file_extension = os.path.splitext(file_path)[1].lower()
        
        # Read the file and load into an appropriate in-memory stream
        if file_extension in ['.xlsx', '.xls']:
            with open(file_path, 'rb') as f:
                in_stream = io.BytesIO(f.read(size))
            file_format = file_extension.lstrip('.')
        elif file_extension == '.csv':
            with open(file_path, 'r', encoding='utf-8') as f:
                in_stream = io.StringIO(f.read(size))
            file_format = 'csv'
        else:
            raise ValueError(f"Unsupported file format: {file_extension}")
        
        return in_stream, file_format


    def filter(self, dataset: tablib.Dataset):
        # Filter out empty rows
        return tablib.Dataset(
            *[row for row in dataset if any(cell not in (None, '', ' ') for cell in row)],
            headers=dataset.headers
        )

    def run(self):
        try:
            batch_size = 500
            total_rows = len(self.dataset)

            for start in range(0, total_rows, batch_size):
                end = min(start + batch_size, total_rows)
                batch = self.dataset[start:end]

                dataset = tablib.Dataset(headers=self.headers)
                dataset.extend(batch)

                self.resource.import_data(
                    dataset=dataset, 
                    raise_errors=self.raise_errors
                )
                
                progress = round((end / total_rows) * 100)
                message = f"Importing records. {progress}% completed."
                self.task.update_state(state="PROGRESS", meta={"progress": progress, "message": message})

            message = f"Task completed. Imported {total_rows} records."
            self.task.update_state(state="SUCCESS", meta={"progress": 100, "message": message})
            if self.notifier:
                self.notifier(self.user, message)
            self.clean_up()

        except ImportError as e:
            error_trace = traceback.format_exc()
            message = f"Error during import: {e}"

            self.task.update_state(
                state="FAILURE",
                meta={
                    "exc_type": type(e).__name__,
                    "exc_message": str(e),
                    "traceback": error_trace,
                    "progress": 100,
                    "message": message
                }
            )
            self.clean_up()

    def clean_up(self):
        """
        Clean up the uploaded file after processing.
        """
        if os.path.exists(self.file_path):
            try:
                # Delete the file from the file system
                os.remove(self.file_path)
            except Exception as e:
                print(f"Error deleting file: {e}")
